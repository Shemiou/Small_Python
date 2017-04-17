#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import time
from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import datetime
import os.path  
import mimetypes
from email.MIMEBase import MIMEBase
from email import Encoders 

import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter 
reload(sys)
sys.setdefaultencoding('utf8')

def send_mail(config):
    print 'Sending Mail...'

    message = MIMEMultipart()
    message["Accept-Charset"] = "ISO-8859-1,utf-8"
    message['From'] = 'pingyang.wang@mail.cn'
    
    message['To'] =','.join(config['to'])
    message['CC'] = ','.join(config['cc'])
    message['Subject'] = config['subject']
    message['Date'] = time.ctime(time.time())
    message['Reply-To'] = 'pingyang.wang@mail.cn'
    message['X-Priority'] = '3'
    message['X-MSMail-Priority'] = 'Normal'
    if config['text']:
        text = config['text']
        message.attach(text)

    part = MIMEApplication(open(fileName,'rb').read())  
    part.add_header('Content-Disposition', 'attachment', filename=fileName)  
    message.attach(part)

    smtp = SMTP(config['server'], config['port'])

    username = 'pingyang.wang@mail.cn'
    smtp.login(username, 'xxxxxx')

    smtp.sendmail(username, config['to'], message.as_string())
    print 'Send Mail OK'

    smtp.close()
    time.sleep(1)

def send_mail_to_test(context):
    send_mail({
        'to': ["wangpingyang03@mail.cn"],
        'cc': ['wangpingyang03@mail.cn'],
        'server': 'smtp.exmail.qq.com',
        'port': 25,
        'subject': 'Just for Test',
        'username': 'pingyang.wang@mail.cn',
        'password': 'xxxxxx',
        'text': context}
    )

def message_from_excel():
    wb = load_workbook(fileName,data_only=True)
    ws = wb.get_sheet_by_name('Crash')

    all_versions = []
    personNums = []
    hanppends = []

    today_bugly = []
    yes_bugly = []

    bugly_flu = []

    for rown in xrange(3,7):
        for coln in xrange(2,8):
            value = ws.cell(row=rown,column=coln).value
            if coln == 2:
                all_versions.append(value)
            elif coln == 3:
                personNums.append(int(value))
            elif coln == 4:
                hanppends.append(int(value))
            elif coln == 5:
                today_bugly.append(float(value))
            elif coln == 6:
                yes_bugly.append(float(value))
            

    for x in xrange(0,4):  
        bugly_flu.append(crash_rate(today_bugly[x],yes_bugly[x]))

    html = """\
<!DOCTYPE html>
<html>
<meta charset="utf-8">
<head>
    <title>iOS - Bugly崩溃日报</title>
</head>
<body>
<div id="container">
    <div id="content">
        <p>

            版本崩溃信息：
            <table width="800" border="2" bordercolor="black" cellspacing="2">
                <tr>
                    <td><strong>版本号</strong></td>
                    <td><strong>影响人数</strong></td>
                    <td><strong>发生次数</strong></td>
                    <td><strong>日崩溃率-用户指标</strong></td>
                    <td><strong>波动</strong></td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[0]) + """</td>
                    <td>""" + str(personNums[0]) + """</td>
                    <td>""" + str(hanppends[0]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[0]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[0]) + """</td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[1]) + """</td>
                    <td>""" + str(personNums[1]) + """</td>
                    <td>""" + str(hanppends[1]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[1]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[1]) + """</td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[2]) + """</td>
                    <td>""" + str(personNums[2]) + """</td>
                    <td>""" + str(hanppends[2]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[2]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[2]) + """</td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[3]) + """</td>
                    <td>""" + str(personNums[3]) + """</td>
                    <td>""" + str(hanppends[3]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[3]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[3]) + """</td>
                </tr>
            </table>
        </p>
        <p>

        详情请见附件

        </p>
    </div>
</div>
</body>
</html>
    """

    context = MIMEText(html,_subtype='html',_charset='utf-8')
    send_mail_to_test(context) 

def crash_rate(today,yester):  
    rate = "%.2f" %(float(today) - float(yester))
    return rate + '%'

def daily_crash_bugly(num):
    temp = "%.2f" %(num * 100) + '%'
    return temp

fileName = '/Users/wangpingyang/Downloads/Bugly-Daily-iOS.xlsx'
print fileName

if __name__ == '__main__':
    message_from_excel()
